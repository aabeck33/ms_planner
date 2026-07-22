import pandas as pd
from datetime import date, datetime, timedelta
import calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from utils import converter_data


def criar_aba_gantt(
    arquivo_saida: str,
    nome_aba_detalhes: str = "Detalhes",
) -> None:
    """
    Cria uma aba Gantt baseada na aba de detalhes.

    Regras:
    - Remove a aba Gantt caso ela já exista.
    - Cria colunas semanais entre a menor Data de início e a maior Data de conclusão.
    - Pinta em azul as semanas compreendidas entre início e fim da ação.
    """

    wb = load_workbook(arquivo_saida)

    if nome_aba_detalhes not in wb.sheetnames:
        raise ValueError(
            f"A aba '{nome_aba_detalhes}' não existe."
        )

    if "Gantt" in wb.sheetnames:
        wb.remove(wb["Gantt"])

    ws_detalhes = wb[nome_aba_detalhes]
    ws_gantt = wb.create_sheet("Gantt")

    dados = list(ws_detalhes.values)

    if len(dados) <= 1:
        wb.save(arquivo_saida)
        return

    cabecalho = list(dados[0])
    linhas = dados[1:]
    
    idx_acao = cabecalho.index("Ação")
    idx_responsavel = cabecalho.index("Responsável")
    idx_inicio = cabecalho.index("Data de início")
    idx_fim = cabecalho.index("Data de conclusão")

    datas_inicio = []
    datas_fim = []

    for linha in linhas:

        acao = linha[idx_acao]
        responsavel = linha[idx_responsavel]

        data_inicio = converter_data(
            linha[idx_inicio]
        )

        data_fim = converter_data(
            linha[idx_fim]
        )

        if not data_inicio or not data_fim:
            continue

        datas_inicio.append(data_inicio)
        datas_fim.append(data_fim)


    if not datas_inicio or not datas_fim:
        wb.save(arquivo_saida)
        return
    
    data_min = min(datas_inicio)
    data_max = max(datas_fim)

    primeira_data = date(
        data_min.year,
        data_min.month,
        1,
    )

    ultimo_dia_mes_final = calendar.monthrange(
        data_max.year,
        data_max.month,
    )[1]

    ultima_data = date(
        data_max.year,
        data_max.month,
        ultimo_dia_mes_final,
    )

    meses_portugues = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }

    colunas_semanais = []

    ano = primeira_data.year
    mes = primeira_data.month

    while True:

        qtd_semanas = len(
            calendar.monthcalendar(ano, mes)
        )

        for semana in range(1, qtd_semanas + 1):

            mes_nome = meses_portugues[mes]

            titulo = (
                f"{mes_nome}/{str(ano)[-2:]} "
                f"S{semana}"
            )

            inicio_semana = date(
                ano,
                mes,
                1
            ) + timedelta(days=(semana - 1) * 7)

            fim_semana = min(
                inicio_semana + timedelta(days=6),
                date(
                    ano,
                    mes,
                    calendar.monthrange(
                        ano,
                        mes
                    )[1]
                ),
            )

            colunas_semanais.append(
                {
                    "titulo": titulo,
                    "inicio": inicio_semana,
                    "fim": fim_semana,
                }
            )

        if ano == ultima_data.year and mes == ultima_data.month:
            break

        mes += 1

        if mes > 12:
            mes = 1
            ano += 1

    ws_gantt.cell(
        row=1,
        column=1,
        value="Ação",
    )

    ws_gantt.cell(
        row=1,
        column=2,
        value="Responsável",
    )

    for coluna, semana in enumerate(
        colunas_semanais,
        start=3,
    ):
        celula = ws_gantt.cell(
            row=1,
            column=coluna,
            value=semana["titulo"],
        )

        celula.font = Font(
            bold=True
        )

        celula.alignment = Alignment(
            horizontal="center"
        )

    azul = PatternFill(
        fill_type="solid",
        fgColor="4F81BD",
    )

    linha_saida = 2

    for linha in linhas:

        acao = linha[idx_acao]
        responsavel = linha[idx_responsavel]
        data_inicio = converter_data(
            linha[idx_inicio]
        )
        data_fim = converter_data(
            linha[idx_fim]
        )

        '''if not data_inicio or not data_fim:
            continue'''
        tem_datas = (
            data_inicio is not None
            and data_fim is not None
        )

        ws_gantt.cell(
            row=linha_saida,
            column=1,
            value=acao,
        )

        ws_gantt.cell(
            row=linha_saida,
            column=2,
            value=responsavel,
        )

        if tem_datas:
            for coluna, semana in enumerate(
                colunas_semanais,
                start=3,
            ):

                if (
                    semana["inicio"] <= data_fim
                    and semana["fim"] >= data_inicio
                ):
                    ws_gantt.cell(
                        row=linha_saida,
                        column=coluna,
                        value=""
                    ).fill = azul

        linha_saida += 1

    ws_gantt.freeze_panes = "C2"

    ws_gantt.column_dimensions["A"].width = 60
    ws_gantt.column_dimensions["B"].width = 25

    for coluna in range(
        3,
        len(colunas_semanais) + 3,
    ):
        letra = ws_gantt.cell(
            row=1,
            column=coluna,
        ).column_letter

        ws_gantt.column_dimensions[
            letra
        ].width = 5

    wb.save(arquivo_saida)
